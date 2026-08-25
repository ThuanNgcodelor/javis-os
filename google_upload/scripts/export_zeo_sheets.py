#!/usr/bin/env python3
"""
Export script: ZeoN8n.xlsx → CSV files cho chatbot
Chạy: python google_upload/scripts/export_zeo_sheets.py

Output:
  google_upload/zeo_faq_google_sheet_from_ZeoN8n_2026_08_13.csv  → rag_search.py dùng
  google_upload/zeo_shopee_catalog_template.csv                   → shopee_matcher.py dùng
"""
import csv
import sys
from pathlib import Path
from urllib.parse import unquote

try:
    import openpyxl
except ImportError:
    print("Missing openpyxl. Run: pip3 install openpyxl")
    sys.exit(1)

BASE = Path(__file__).resolve().parents[1]
XLSX = BASE / "ZeoN8n.xlsx"

OUTPUTS = {
    "FAQ":    BASE / "zeo_faq_google_sheet_from_ZeoN8n_2026_08_13.csv",
    "Shopee": BASE / "zeo_shopee_catalog_template.csv",
}

def export_sheet(ws, out_path: Path, decode_url_col: int = None):
    rows_written = 0
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            row = list(row)
            if decode_url_col is not None and len(row) > decode_url_col and row[decode_url_col]:
                row[decode_url_col] = unquote(str(row[decode_url_col]))
            writer.writerow(row)
            if i > 0:
                rows_written += 1
    return rows_written

def main():
    if not XLSX.exists():
        print(f"File not found: {XLSX}")
        sys.exit(1)

    wb = openpyxl.load_workbook(XLSX)
    print(f"Loaded: {XLSX.name}")
    print(f"Sheets: {wb.sheetnames}")

    if "FAQ" in wb.sheetnames:
        n = export_sheet(wb["FAQ"], OUTPUTS["FAQ"])
        print(f"FAQ -> {OUTPUTS['FAQ'].name} ({n} rows)")
    else:
        print("Sheet 'FAQ' not found, skipping")

    if "Shopee" in wb.sheetnames:
        n = export_sheet(wb["Shopee"], OUTPUTS["Shopee"], decode_url_col=12)
        print(f"Shopee -> {OUTPUTS['Shopee'].name} ({n} rows)")
    else:
        print("Sheet 'Shopee' not found, skipping")

    print("Export done! Chatbot dung CSV moi ngay khi restart hoac /sync")

if __name__ == "__main__":
    main()
