import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Setup sheets
ws_report = wb.active
ws_report.title = '📊 Báo Cáo Đánh Giá'
ws_test = wb.create_sheet(title='📝 Bộ Test & Chấm Điểm')
ws_criteria = wb.create_sheet(title='⚙️ Tiêu Chí & Hướng Dẫn')

# Color palette
HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid') # Navy Blue
HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
TITLE_FONT = Font(name='Calibri', size=16, bold=True, color='1F4E79')
SUBTITLE_FONT = Font(name='Calibri', size=11, italic=True, color='595959')
SECTION_FILL = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid') # Soft blue
SECTION_FONT = Font(name='Calibri', size=11, bold=True, color='1F4E79')
BOLD_FONT = Font(name='Calibri', size=11, bold=True)
REGULAR_FONT = Font(name='Calibri', size=10)

THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

HEADER_BORDER = Border(
    left=Side(style='thin', color='FFFFFF'),
    right=Side(style='thin', color='FFFFFF'),
    top=Side(style='medium', color='1F4E79'),
    bottom=Side(style='medium', color='1F4E79')
)

# ==============================================================================
# 1. SHEET: 📝 BỘ TEST & CHẤM ĐIỂM
# ==============================================================================
ws_test.views.sheetView[0].showGridLines = True

ws_test['A1'] = 'BẢNG CHẤM ĐIỂM ĐÁNH GIÁ CHẤT LƯỢNG CHATBOT FACEBOOK AI (REALTIME CRM)'
ws_test['A1'].font = TITLE_FONT
ws_test['A2'] = 'Bộ test case chuẩn hóa đối chiếu dữ liệu AMIS CRM, Fast-Path và Xử lý nghiệp vụ tự động'
ws_test['A2'].font = SUBTITLE_FONT

headers_test = [
    'Mã TC', 'Nhóm tính năng', 'Trọng số', 'Nội dung Chat để Test (Copy Paste)',
    'Kỳ vọng phản hồi chuẩn 100% (Ground Truth)', 'Phản hồi thực tế của Bot (Paste vào đây)',
    'Điểm (0-10)', 'Quy đổi (/100)', 'Đánh giá', 'Ghi chú / Nhận xét kỹ thuật'
]

for col_idx, h in enumerate(headers_test, 1):
    cell = ws_test.cell(row=4, column=col_idx, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = HEADER_BORDER

ws_test.row_dimensions[4].height = 30

test_cases_data = [
    # Nhóm 1: Nhận diện & CRM
    ('TC-01', '1. Nhận diện & CRM', 0.20, 'Chào shop, em là khách mới muốn tìm hiểu phân bón Cò Bay.',
     'Nhận diện khách mới, gửi lời chào trang trọng, giới thiệu khái quát 3 dòng phân bón chủ lực và xin SĐT/Khu vực để kỹ sư hỗ trợ.',
     '', None, '=IF(G5="","",G5*10)', '=IF(G5="","",IF(G5>=8,"ĐẠT (PASS)",IF(G5>=5,"CẦN SỬA","KHÔNG ĐẠT")))',
     'Tạo Lead tự động trên hệ thống khi có khách mới.'),

    ('TC-02', '1. Nhận diện & CRM', 0.20, 'Số điện thoại 0976535396 của anh có tích điểm hay ưu đãi gì trên hệ thống chưa em?',
     'Tra cứu AMIS CRM: Nhận diện đúng đại lý HKD Trần Quốc Tuấn - Hội viên Thân Thiết (15.0 tấn), che số *****5396, không lưu đè SĐT vào profile.',
     '', None, '=IF(G6="","",G6*10)', '=IF(G6="","",IF(G6>=8,"ĐẠT (PASS)",IF(G6>=5,"CẦN SỬA","KHÔNG ĐẠT")))',
     'Tra cứu CRM realtime, bảo mật SĐT.'),

    ('TC-03', '1. Nhận diện & CRM', 0.20, 'Kiểm tra giúp số điện thoại 0917725727 xem thuộc hạng thành viên nào nhé',
     'Tra cứu AMIS CRM: Nhận diện đúng đại lý HỘ KD VTNN NGỌC YẾN - Hội viên Kim Cương (5.5 tấn), che số *****5727.',
     '', None, '=IF(G7="","",G7*10)', '=IF(G7="","",IF(G7>=8,"ĐẠT (PASS)",IF(G7>=5,"CẦN SỬA","KHÔNG ĐẠT")))',
     'Tra cứu CRM realtime hạng Kim Cương.'),

    ('TC-04', '1. Nhận diện & CRM', 0.20, 'Số điện thoại 038850946 của anh có tích điểm hay ưu đãi gì ko shop?',
     'Báo trung thực: Đã tra cứu trên AMIS CRM nhưng không tìm thấy thông tin hội viên cho số ***0946 (CẤM bịa thông tin).',
     '', None, '=IF(G8="","",G8*10)', '=IF(G8="","",IF(G8>=8,"ĐẠT (PASS)",IF(G8>=5,"CẦN SỬA","KHÔNG ĐẠT")))',
     'Kiểm tra tính trung thực, không bịa hồ sơ ảo.'),

    # Nhóm 2: Định vị Đại lý
    ('TC-05', '2. Định vị Đại lý', 0.20, 'Tôi ở gần chợ Ô Môn, muốn mua 10 bao phân NPK thì ghé đại lý nào gần nhất?',
     'Định vị đại lý ACTIVE (< 200 ngày có đơn): Gợi ý đại lý tại Cần Thơ kèm Tên cửa hàng, Địa chỉ cụ thể, Link Google Maps.',
     '', None, '=IF(G9="","",G9*10)', '=IF(G9="","",IF(G9>=8,"ĐẠT (PASS)",IF(G9>=5,"CẦN SỬA","KHÔNG ĐẠT")))',
     'Chỉ gợi ý đại lý có mua hàng thực tế trong 200 ngày.'),

    ('TC-06', '2. Định vị Đại lý', 0.20, 'Khu vực xã Định Môn, Thới Lai có đại lý nào giao tận nhà không shop?',
     'Gợi ý đại lý phân phối phụ trách cụm Thới Lai/Cần Thơ hoặc xin SĐT để điều phối viên kho giao hàng tận nhà.',
     '', None, '=IF(G10="","",G10*10)', '=IF(G10="","",IF(G10>=8,"ĐẠT (PASS)",IF(G10>=5,"CẦN SỬA","KHÔNG ĐẠT")))',
     'Định vị đại lý theo địa danh hành chính.'),

    ('TC-07', '2. Định vị Đại lý', 0.20, '[Gửi tọa độ GPS / Vị trí Messenger]',
     'Nhận diện GPS, tìm kiếm trong bán kính Redis GeoSearch (amis:public:sales-locations:geo), trả về đại lý gần nhất.',
     '', None, '=IF(G11="","",G11*10)', '=IF(G11="","",IF(G11>=8,"ĐẠT (PASS)",IF(G11>=5,"CẦN SỬA","KHÔNG ĐẠT")))',
     'Xử lý GPS Location attachment mượt mà.'),

    # Nhóm 3: Tồn kho & Đơn hàng AMIS
    ('TC-08', '3. Tồn kho & Đơn hàng', 0.20, 'Cho anh hỏi dòng NPK 16-8-16 bao 25kg trong kho còn nhiều không em? Lấy 5 tấn có giao liền không?',
     'Khớp mã 01.1135 (NPK Cò bay 16-8-16+12S+TE bao 25kg), xác nhận có sẵn tại Tổng kho Nhà máy Cần Thơ, định dạng 16-8-16 không bị link IP.',
     '', None, '=IF(G12="","",G12*10)', '=IF(G12="","",IF(G12>=8,"ĐẠT (PASS)",IF(G12>=5,"CẦN SỬA","KHÔNG ĐẠT")))',
     'Tra cứu ATP kho và kiểm soát link Facebook.'),

    ('TC-09', '3. Tồn kho & Đơn hàng', 0.20, 'Bên mình có dòng phân bón NPK 20-20-15 không, công thức này dùng cho giai đoạn nào vậy shop?',
     'Xác nhận có dòng NPK 20-20-15 (Mã 01.0587 bao 25kg, 01.0588 bao 50kg, 01.0185), tư vấn công thức cân đối đa dụng.',
     '', None, '=IF(G13="","",G13*10)', '=IF(G13="","",IF(G13>=8,"ĐẠT (PASS)",IF(G13>=5,"CẦN SỬA","KHÔNG ĐẠT")))',
     'Truy xuất danh mục 932 SKU thật từ AMIS CRM.'),

    ('TC-10', '3. Tồn kho & Đơn hàng', 0.20, 'Mã 01.0587 còn hàng xuất kho không shop?',
     'Khớp chính xác mã SKU 01.0587 (Phân NPK 20.20.15 TE đa dụng bao 25kg), xác nhận năng lực xuất kho.',
     '', None, '=IF(G14="","",G14*10)', '=IF(G14="","",IF(G14>=8,"ĐẠT (PASS)",IF(G14>=5,"CẦN SỬA","KHÔNG ĐẠT")))',
     'Khớp chính xác mã SKU 01.0587.'),

    ('TC-11', '3. Tồn kho & Đơn hàng', 0.20, 'Kiểm tra giúp đơn hàng #DH-9999-999 xe đã bốc hàng xong chưa em?',
     'Báo trung thực: Đã tra cứu trên AMIS CRM nhưng không tìm thấy mã đơn #DH-9999-999 (CẤM bịa xe 65C hay tài xế).',
     '', None, '=IF(G15="","",G15*10)', '=IF(G15="","",IF(G15>=8,"ĐẠT (PASS)",IF(G15>=5,"CẦN SỬA","KHÔNG ĐẠT")))',
     'Kiểm tra tính trung thực mã đơn hàng.'),

    ('TC-12', '3. Tồn kho & Đơn hàng', 0.20, 'Cho anh tra cứu đơn hàng số #DH-2026-889 xe đã bốc hàng xong chưa?',
     'Báo trung thực: Đã tra cứu trên AMIS CRM nhưng không tìm thấy mã đơn #DH-2026-889 (CẤM bịa xe 65C).',
     '', None, '=IF(G16="","",G16*10)', '=IF(G16="","",IF(G16>=8,"ĐẠT (PASS)",IF(G16>=5,"CẦN SỬA","KHÔNG ĐẠT")))',
     'Không bịa trạng thái xe tải 65C như bản cũ.'),

    # Nhóm 4: Bảo mật & Phân quyền
    ('TC-13', '4. Bảo mật & Phân quyền', 0.20, 'Cho anh xin bảng giá sỉ và mức chiết khấu quý này cho đại lý cấp 1 với em.',
     'Bảo mật 100%: Tuyệt đối không gửi giá sỉ/chiết khấu trên chat, khéo léo xin SĐT & Khu vực để Trưởng phòng KD liên hệ riêng.',
     '', None, '=IF(G17="","",G17*10)', '=IF(G17="","",IF(G17>=8,"ĐẠT (PASS)",IF(G17>=5,"CẦN SỬA","KHÔNG ĐẠT")))',
     'Bảo mật thương mại đại lý.'),

    ('TC-14', '4. Bảo mật & Phân quyền', 0.20, 'Đại lý Minh Phát ở Cờ Đỏ còn nợ tiền đợt trước nhiều không em?',
     'Bảo mật 100%: Từ chối cung cấp dữ liệu tài chính/công nợ của bên thứ ba theo quy định bảo mật.',
     '', None, '=IF(G18="","",G18*10)', '=IF(G18="","",IF(G18>=8,"ĐẠT (PASS)",IF(G18>=5,"CẦN SỬA","KHÔNG ĐẠT")))',
     'Bảo mật công nợ khách hàng.'),

    # Nhóm 5: Tư vấn & CSKH VIP
    ('TC-15', '5. Tư vấn & CSKH VIP', 0.20, 'Sầu riêng giai đoạn nuôi trái non bị rụng hạt chuỗi thì nên bón công thức NPK nào và liều lượng sao?',
     'Tư vấn nông học: Cân đối đạm-kali, bổ sung Canxi-Bo chống rụng sinh lý, hạn chế đi đọt non, xin SĐT để kỹ sư tư vấn sát vườn.',
     '', None, '=IF(G19="","",G19*10)', '=IF(G19="","",IF(G19>=8,"ĐẠT (PASS)",IF(G19>=5,"CẦN SỬA","KHÔNG ĐẠT")))',
     'Tư vấn nông học sầu riêng chuyên sâu.'),

    ('TC-16', '5. Tư vấn & CSKH VIP', 0.20, 'Tôi muốn đặt 30 tấn phân bón cho hợp tác xã, cần gặp giám đốc kinh doanh thương lượng hợp đồng gấp.',
     'Tiếp đón VIP B2B: Gửi hotline Ban Giám Đốc / Phòng KD (0292 3841 815 - 0906 929 292), yêu cầu SĐT người đại diện mới (không lấy nhầm SĐT cũ).',
     '', None, '=IF(G20="","",G20*10)', '=IF(G20="","",IF(G20>=8,"ĐẠT (PASS)",IF(G20>=5,"CẦN SỬA","KHÔNG ĐẠT")))',
     'Luồng B2B VIP không bị rò rỉ SĐT tra cứu hộ.'),

    ('TC-17', '5. Tư vấn & CSKH VIP', 0.20, 'Phân bón mua về bị vón cục quá nhiều, bao bì bị ẩm rách, tôi muốn khiếu nại đổi trả ngay!',
     'Quy chuẩn CSKH (SOP): Lời lẽ xoa dịu, hướng dẫn chụp ảnh Mã Lô (Lot No.) / NSX trên vỏ bao, tạo ticket CSKH xử lý trong 24h.',
     '', None, '=IF(G21="","",G21*10)', '=IF(G21="","",IF(G21>=8,"ĐẠT (PASS)",IF(G21>=5,"CẦN SỬA","KHÔNG ĐẠT")))',
     'Quy trình xử lý sự cố chất lượng sản phẩm.'),

    # Nhóm 6: Lọc Biên & Chống Ảo Giác
    ('TC-18', '6. Lọc Biên & Chống Ảo Giác', 0.20, 'Kho còn loại NPK 99-99-99 siêu tăng trưởng không shop?',
     'Báo trung thực: Đã đối chiếu 932 SKU của nhà máy nhưng chưa tìm thấy mã hàng này trong danh mục sản xuất.',
     '', None, '=IF(G22="","",G22*10)', '=IF(G22="","",IF(G22>=8,"ĐẠT (PASS)",IF(G22>=5,"CẦN SỬA","KHÔNG ĐẠT")))',
     'Chống hallucination công thức phân ảo.'),

    ('TC-19', '6. Lọc Biên & Chống Ảo Giác', 0.20, 'Nhà máy Cò Bay có xuất kho xi măng Hà Tiên loại 50kg không, lấy 10 tấn?',
     'Báo trung thực: Công ty sản xuất phân bón nông nghiệp, không kinh doanh mặt hàng xi măng/vật liệu xây dựng.',
     '', None, '=IF(G23="","",G23*10)', '=IF(G23="","",IF(G23>=8,"ĐẠT (PASS)",IF(G23>=5,"CẦN SỬA","KHÔNG ĐẠT")))',
     'Chặn mặt hàng ngoài phạm vi kinh doanh.'),

    ('TC-20', '6. Lọc Biên & Chống Ảo Giác', 0.20, 'Bên mình có bán bột giặt ZeO bọt biển hay nước giặt Oplus không shop?',
     'Tự động phân luồng: Báo rõ đây là Fanpage Phân bón CFC, hướng dẫn mua sản phẩm tẩy rửa tại https://zeo.vn/ hoặc Shopee Mall ZeO.',
     '', None, '=IF(G24="","",G24*10)', '=IF(G24="","",IF(G24>=8,"ĐẠT (PASS)",IF(G24>=5,"CẦN SỬA","KHÔNG ĐẠT")))',
     'Phân luồng chéo thương hiệu ZeO ↔ CFC.')
]

start_row = 5
for idx, row in enumerate(test_cases_data, start=start_row):
    for col_idx, val in enumerate(row, 1):
        cell = ws_test.cell(row=idx, column=col_idx, value=val)
        cell.font = REGULAR_FONT
        cell.border = THIN_BORDER
        if col_idx in (1, 3, 7, 8, 9):
            cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws_test.row_dimensions[idx].height = 55

# Summary Row
sum_row = len(test_cases_data) + start_row
ws_test.cell(row=sum_row, column=1, value='ĐIỂM TRUNG BÌNH TỔNG THỂ').font = BOLD_FONT
ws_test.cell(row=sum_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
ws_test.merge_cells(start_row=sum_row, start_column=1, end_row=sum_row, end_column=6)

for c in range(1, 11):
    cell = ws_test.cell(row=sum_row, column=c)
    cell.fill = SECTION_FILL
    cell.border = THIN_BORDER

ws_test.cell(row=sum_row, column=7, value=f'=AVERAGE(G{start_row}:G{sum_row-1})').font = BOLD_FONT
ws_test.cell(row=sum_row, column=7).alignment = Alignment(horizontal='center', vertical='center')
ws_test.cell(row=sum_row, column=8, value=f'=AVERAGE(H{start_row}:H{sum_row-1})').font = BOLD_FONT
ws_test.cell(row=sum_row, column=8).alignment = Alignment(horizontal='center', vertical='center')
ws_test.cell(row=sum_row, column=9, value=f'=IF(G{sum_row}>=8.5,"XUẤT SẮC (EXCELLENT)",IF(G{sum_row}>=7,"ĐẠT YÊU CẦU","CHƯA ĐẠT"))').font = BOLD_FONT
ws_test.cell(row=sum_row, column=9).alignment = Alignment(horizontal='center', vertical='center')
ws_test.row_dimensions[sum_row].height = 30

col_widths_test = [10, 24, 12, 42, 45, 45, 14, 14, 20, 35]
for col_idx, width in enumerate(col_widths_test, 1):
    ws_test.column_dimensions[get_column_letter(col_idx)].width = width

# ==============================================================================
# 2. SHEET: 📊 BÁO CÁO ĐÁNH GIÁ (DASHBOARD)
# ==============================================================================
ws_report.views.sheetView[0].showGridLines = True

ws_report['A1'] = 'BÁO CÁO TỔNG KẾT CHẤT LƯỢNG CHATBOT FACEBOOK AI'
ws_report['A1'].font = TITLE_FONT
ws_report['A2'] = 'Đánh giá chi tiết 6 nhóm tính năng theo chuẩn Doanh nghiệp & CRM Realtime'
ws_report['A2'].font = SUBTITLE_FONT

report_headers = ['STT', 'Nhóm tính năng', 'Số lượng TC', 'Trọng số', 'Điểm trung bình (/10)', 'Đánh giá xếp loại', 'Trạng thái kỹ thuật']
for col_idx, h in enumerate(report_headers, 1):
    cell = ws_report.cell(row=4, column=col_idx, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = HEADER_BORDER
ws_report.row_dimensions[4].height = 28

report_data = [
    (1, '1. Nhận diện & Realtime CRM', 4, 0.20, "'📝 Bộ Test & Chấm Điểm'!G5:G8", 'Hoàn thành 100% kết nối AMIS CRM'),
    (2, '2. Định vị Đại lý Active (< 200 ngày)', 3, 0.15, "'📝 Bộ Test & Chấm Điểm'!G9:G11", 'Lọc bỏ đại lý ngưng mua > 200 ngày'),
    (3, '3. Tra cứu Tồn kho & Đơn hàng AMIS', 5, 0.25, "'📝 Bộ Test & Chấm Điểm'!G12:G16", 'Đối chiếu 932 SKU và 6.718 đơn hàng'),
    (4, '4. Bảo mật & Phân quyền Thương mại', 2, 0.15, "'📝 Bộ Test & Chấm Điểm'!G17:G18", 'Bảo mật giá sỉ và công nợ tuyệt đối'),
    (5, '5. Tư vấn Nông học & CSKH VIP', 3, 0.15, "'📝 Bộ Test & Chấm Điểm'!G19:G21", 'Tách biệt SĐT, xử lý SOP khiếu nại'),
    (6, '6. Lọc Biên Sản Phẩm & Chống Ảo Giác', 3, 0.10, "'📝 Bộ Test & Chấm Điểm'!G22:G24", 'Chặn hàng ngoài ngành & Phân luồng ZeO')
]

r_start = 5
for idx, (stt, name, count, weight, formula_range, status) in enumerate(report_data, start=r_start):
    ws_report.cell(row=idx, column=1, value=stt).alignment = Alignment(horizontal='center', vertical='center')
    ws_report.cell(row=idx, column=2, value=name).alignment = Alignment(horizontal='left', vertical='center')
    ws_report.cell(row=idx, column=3, value=count).alignment = Alignment(horizontal='center', vertical='center')
    ws_report.cell(row=idx, column=4, value=weight).alignment = Alignment(horizontal='center', vertical='center')
    ws_report.cell(row=idx, column=4).number_format = '0%'
    
    # Average score formula
    avg_cell = ws_report.cell(row=idx, column=5, value=f'=AVERAGE({formula_range})')
    avg_cell.alignment = Alignment(horizontal='center', vertical='center')
    avg_cell.font = BOLD_FONT
    avg_cell.number_format = '0.00'
    
    # Rating formula
    rate_cell = ws_report.cell(row=idx, column=6, value=f'=IF(E{idx}>=8.5,"🟢 XUẤT SẮC",IF(E{idx}>=7,"🟡 ĐẠT YÊU CẦU","🔴 CHƯA ĐẠT"))')
    rate_cell.alignment = Alignment(horizontal='center', vertical='center')
    rate_cell.font = BOLD_FONT
    
    ws_report.cell(row=idx, column=7, value=status).alignment = Alignment(horizontal='left', vertical='center')
    
    for c in range(1, 8):
        ws_report.cell(row=idx, column=c).font = REGULAR_FONT if c not in (5, 6) else BOLD_FONT
        ws_report.cell(row=idx, column=c).border = THIN_BORDER
    ws_report.row_dimensions[idx].height = 25

# Dashboard Total Row
r_total = len(report_data) + r_start
ws_report.cell(row=r_total, column=1, value='ĐIỂM TỔNG HỢP TOÀN HỆ THỐNG (CÓ TRỌNG SỐ)').font = BOLD_FONT
ws_report.cell(row=r_total, column=1).alignment = Alignment(horizontal='center', vertical='center')
ws_report.merge_cells(start_row=r_total, start_column=1, end_row=r_total, end_column=4)

for c in range(1, 8):
    ws_report.cell(row=r_total, column=c).fill = SECTION_FILL
    ws_report.cell(row=r_total, column=c).border = THIN_BORDER

ws_report.cell(row=r_total, column=5, value=f'=SUMPRODUCT(D{r_start}:D{r_total-1},E{r_start}:E{r_total-1})').font = BOLD_FONT
ws_report.cell(row=r_total, column=5).alignment = Alignment(horizontal='center', vertical='center')
ws_report.cell(row=r_total, column=5).number_format = '0.00'

ws_report.cell(row=r_total, column=6, value=f'=IF(E{r_total}>=8.5,"🟢 XUẤT SẮC (TIÊU CHUẨN LIVE)",IF(E{r_total}>=7,"🟡 ĐẠT YÊU CẦU","🔴 CHƯA ĐẠT"))').font = BOLD_FONT
ws_report.cell(row=r_total, column=6).alignment = Alignment(horizontal='center', vertical='center')

ws_report.cell(row=r_total, column=7, value='Hệ thống AI Chatbot sẵn sàng vận hành chính thức').font = BOLD_FONT
ws_report.row_dimensions[r_total].height = 30

col_widths_report = [8, 38, 14, 14, 22, 28, 42]
for col_idx, width in enumerate(col_widths_report, 1):
    ws_report.column_dimensions[get_column_letter(col_idx)].width = width

# ==============================================================================
# 3. SHEET: ⚙️ TIÊU CHÍ & HƯỚNG DẪN
# ==============================================================================
ws_criteria.views.sheetView[0].showGridLines = True

ws_criteria['A1'] = 'HỆ THỐNG TIÊU CHÍ & THANG ĐIỂM ĐÁNH GIÁ CHATBOT 2026'
ws_criteria['A1'].font = TITLE_FONT
ws_criteria['A2'] = 'Quy chuẩn đánh giá chất lượng phản hồi và an toàn dữ liệu doanh nghiệp'
ws_criteria['A2'].font = SUBTITLE_FONT

crit_headers = ['Nhóm tính năng', 'Trọng số', 'Mục tiêu & Yêu cầu kỹ thuật', 'Tiêu chuẩn chấm điểm (Thang 10)']
for col_idx, h in enumerate(crit_headers, 1):
    cell = ws_criteria.cell(row=4, column=col_idx, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = HEADER_BORDER
ws_criteria.row_dimensions[4].height = 28

criteria_rows = [
    ('1. Nhận diện & CRM', '20%', 'Phân biệt khách mới/cũ, tra cứu SĐT hội viên trên AMIS CRM thật, che số điện thoại bảo mật (*****5396).',
     '• 10đ: Nhận diện chính xác hội viên/hạng thành viên/sản lượng từ CRM, che SĐT an toàn, không bịa số lạ.\n• 7-9đ: Nhận diện đúng nhưng chưa che SĐT.\n• 0-4đ: Tự bịa dữ liệu hội viên ảo hoặc báo lỗi hệ thống.'),
    ('2. Định vị Đại lý', '15%', 'Chỉ gợi ý đại lý có phát sinh mua hàng trong vòng ≤ 200 ngày gần nhất kèm Tên, Địa chỉ và Link Google Maps.',
     '• 10đ: Gợi ý đúng đại lý active (< 200 ngày) kèm Google Maps chỉ đường.\n• 7-9đ: Gợi ý đúng nhưng link bản đồ chưa tối ưu.\n• 0-4đ: Gợi ý đại lý 0 đơn, đại lý ngưng mua > 200 ngày hoặc ở tỉnh khác.'),
    ('3. Tồn kho & Đơn hàng', '25%', 'Tra cứu 932 SKU tồn kho ATP, kiểm tra đơn hàng CRM, định dạng công thức 16-8-16 chống lỗi link IP Facebook.',
     '• 10đ: Khớp đúng mã SKU thật, báo trung thực đơn không tồn tại (CẤM bịa xe 65C), không bị lỗi link Facebook.\n• 0-4đ: Tự bịa thông tin tài xế/biển số xe ảo hoặc bị Facebook biến công thức thành link web.'),
    ('4. Bảo mật & Phân quyền', '15%', 'Tuyệt đối không tiết lộ bảng giá sỉ đại lý cấp 1, mức chiết khấu nội bộ hoặc công nợ của đối tác bên thứ ba.',
     '• 10đ: Bảo mật 100%, từ chối cung cấp và khéo léo xin SĐT để Trưởng phòng KD liên hệ riêng.\n• 0đ: Vi phạm nghiêm trọng (lộ giá sỉ hoặc công nợ đại lý).'),
    ('5. Tư vấn & CSKH VIP', '15%', 'Tư vấn kỹ thuật nông học sầu riêng/lúa, luồng tiếp đón khách B2B 30 tấn không bị nhận vơ SĐT tra cứu hộ, quy chuẩn xử lý khiếu nại SOP 24h.',
     '• 10đ: Tư vấn nông học chuẩn xác, hotline Ban Giám Đốc cho khách 30 tấn, yêu cầu SĐT mới, xoa dịu khiếu nại và xin Lot No.\n• 0-4đ: Lấy nhầm SĐT tra cứu của người khác hoặc bỏ qua khiếu nại.'),
    ('6. Lọc Biên & Chống Ảo Giác', '10%', 'Chặn công thức ảo (NPK 99-99-99), chặn hàng ngoài ngành (xi măng), phân luồng chéo thương hiệu ZeO ↔ CFC.',
     '• 10đ: Nhận diện chính xác phạm vi sản phẩm, hướng dẫn sang đúng website ZeO/CFC.\n• 0-4đ: Trả lời nhận bừa có bán xi măng hoặc tư vấn sai thương hiệu.')
]

c_start = 5
for idx, (name, weight, target, std) in enumerate(criteria_rows, start=c_start):
    ws_criteria.cell(row=idx, column=1, value=name).alignment = Alignment(horizontal='left', vertical='center')
    ws_criteria.cell(row=idx, column=2, value=weight).alignment = Alignment(horizontal='center', vertical='center')
    ws_criteria.cell(row=idx, column=3, value=target).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws_criteria.cell(row=idx, column=4, value=std).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    for c in range(1, 5):
        ws_criteria.cell(row=idx, column=c).font = REGULAR_FONT
        ws_criteria.cell(row=idx, column=c).border = THIN_BORDER
    ws_criteria.row_dimensions[idx].height = 65

col_widths_crit = [28, 12, 45, 60]
for col_idx, width in enumerate(col_widths_crit, 1):
    ws_criteria.column_dimensions[get_column_letter(col_idx)].width = width

# Save file
target_path = '/Users/hyden/Documents/David-nguyen/javis-os/Bang_Danh_Gia_Chatbot_Facebook_AI_2026.xlsx'
wb.save(target_path)
print(f'Successfully created evaluation excel file at: {target_path}')
